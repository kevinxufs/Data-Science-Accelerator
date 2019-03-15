import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from pandas import ExcelWriter
from pandas import ExcelFile



#Will need to set a source directory that contains all the files contained in the 'datasets' list below.

datasets = ['sep17.xlsx',
 'hs2_airquality_data_areasouth_may_2018.xlsx',
 'hs2_airquality_data_march_2018.xlsx',
 'hs2_airquality_data_areasouth_july_2018.xlsx',
 'oct17.xlsx',
 'dec17.xlsx',
 'nov17.xlsx',
 'hs2_airquality_jan_2018.xlsx',
 'hs2_airquality_data_april_2018.xlsx',
 'hs2_airquality_data_areasouth_june_2018.xlsx',
 'hs2_airquality_data_february_2018.xlsx']

os.chdir('/Users/datascience9/Desktop/air quality')


#metarefdata is an excel file that contains information on all of the sensor locations.

metarefdata = load_workbook('metarefdata.xlsx')


x = metarefdata.active
all_rows = []

for row in x:
    current_row = []
    for cell in row:
        current_row.append(cell.value)
    all_rows.append(current_row)
    
completedic = {}

#Use the metaref data to create a dictionary with the first value (Monitor ref) as a key and the latter values in a list for each dictionary key
for lists in all_rows:
    completedic[lists[0]] = lists[1:]
    
    
    
#Create an empty dataframe with the following column names. This dataframe will be filled to produce an excel and csv file of
#all of the air quality data

col = ['Date/Time', 'PM10 particles (ug/m^3)', 'Monitor Ref', 'Location 2', 'London Borough', 'GPS reference 1', 'GPS reference 2', 'Grid reference', 'Latitude', 'Longitude', 'Postcode', 'Worksite Ref']

completedf = pd.DataFrame(columns = col)

#For each file we load the workbook and then iterate through the sheets. For each sheet, we read the dataset and extract
#the monitor ref from it. The monitor ref is the required dictionary key from our metaref data. We use this dictionary key to 
#fill out the remaining columns.
#After each sheet is processed, it is appended to our empty dataframe. The end result is a combined dataframe.

for file in datasets:
    wb = load_workbook(file)
    sheets = wb.sheetnames[1:len(wb.sheetnames)]
    print('Now processing file ' + file)
    for i in sheets:
        df = pd.read_excel(file, sheet_name = i, skiprows = 3)
        
        #the monitor reference is located at the 'B1' location on each worksheet
        ref = wb[i]['B1'].value
        
        
        df['Monitor Ref'] = ref
        df['Location 2'] = completedic[ref][0]
        df['London Borough'] = completedic[ref][1]
        df['GPS reference 1'] = completedic[ref][2]
        df['GPS reference 2'] = completedic[ref][3]
        df['Grid reference'] = completedic[ref][4]
        df['Latitude'] = completedic[ref][5]
        df['Longitude'] = completedic[ref][6]
        df['Postcode'] = completedic[ref][7]
        df['Worksite Ref'] = completedic[ref][8]
        completedf = pd.concat([completedf, df])
        
        
#Save the data as an excel file. This can also be converted into a csv file.

writer = pd.ExcelWriter('completedatav2.xlsx', engine ='xlsxwriter')

completedf.to_excel(writer, sheet_name = 'Sheet1', index = False)

writer.save()
        
